from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

path = "C:\\Users\\stressunfit\\Documents\\xlpractice\\Book2.xlsx"

workbook = load_workbook(path)
worksheet = workbook.active

col = worksheet.max_column
row = worksheet.max_row
print(row)
print(col)
for r in range(1, row+1):
    for c in range(1, col+1):
        # char = get_column_letter(col)
        print(worksheet.cell(row = r, column=c).value,end = "    ")

    print()
